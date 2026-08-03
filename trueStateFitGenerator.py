import pandas as pd
import openpyxl
import os
from tqdm import tqdm
from CRD_DB_Lookup import lookup_crd
import datetime

def get_unique_filename(file_path):
    """Checks if a file exists and appends a numeric suffix if it does."""
    if not os.path.exists(file_path):
        return file_path

    # Split into file path/name and the .xlsx extension
    base, extension = os.path.splitext(file_path)
    counter = 1
    
    # Try 'FileName 1.xlsx', 'FileName 2.xlsx', etc.
    new_path = f"{base} {counter}{extension}"
    while os.path.exists(new_path):
        counter += 1
        new_path = f"{base} {counter}{extension}"
        
    return new_path

def cleanName(firstname: str, lastname: str) -> str:
    if firstname and lastname:
        return f"{firstname.split()[0].capitalize()} {lastname.split()[0].capitalize()}"
    return None

def addTrueState(fitPath: str, oldFitPath: str, fitSheet: str = "FIT", oldFitSheet: str = "FIT"):
    '''
    Makes two passes through fit list - one building a mapping of names to registrations (correcting for inconsistent CRD info), and one applying the correct state based on the mapping and rules.
    '''
    old = pd.read_excel(oldFitPath, sheet_name=oldFitSheet, header=1)
    oldDF = pd.DataFrame(old)
    existingStates = {
        cleanName(first, last): state
        for first, last, state in zip(oldDF['First'], oldDF['Last'], oldDF['Home State'])
        if cleanName(first, last) is not None
    }

    advisors = {}
    wb = openpyxl.load_workbook(fitPath)
    ws = wb[fitSheet]
    
    target_index = 10 # Column J - Inserting 2 columns before it (pushing existing content back)
    cols_to_insert = 2
    ws.insert_cols(target_index, cols_to_insert)
    ws.cell(row=2, column=target_index).value = "Home State"
    ws.cell(row=2, column=target_index + 1).value = "Full Name"
    
    # LOOP 1: Build the dictionary
    for row in tqdm(ws.iter_rows(min_row=3, min_col=10, max_col=18), desc="Processing names"): 
        cell_fullname = row[1]  
        cell_CRD = row[2]  
        cell_lastname = row[7]  
        cell_firstname = row[8]  
        cell_fullname.value = cleanName(cell_firstname.value, cell_lastname.value)

        # Handle missing CRDs safely
        if not cell_CRD.value:
            continue
            
        try:
            crd_info = lookup_crd(int(cell_CRD.value))
        except ValueError:
            continue

        if crd_info is None or cell_fullname.value != cleanName(crd_info.get("first_name"), crd_info.get("last_name")):
            continue
            
        # FIX: Store just the registrations dict
        advisors[cell_fullname.value] = crd_info.get('registrations', {})
    
    # LOOP 2: Apply logic
    for row in tqdm(ws.iter_rows(min_row=3, min_col=10, max_col=18), desc="Processing rows"): 
        cell_homestate = row[0]  
        cell_fullname = row[1]  
        cell_CRD = row[2]  
        cell_lastname = row[7]  
        cell_firstname = row[8]  
        cell_date = row[3]  
        
        regs = advisors.get(cell_fullname.value, {})
        all_states = [state for states_list in regs.values() for state in states_list]
        old_state = existingStates.get(cell_fullname.value)

        # --- PREPARATION: Format Dates for Priority 1 ---
        # 1. Format the Excel cell date
        if hasattr(cell_date.value, 'strftime'):
            fit_list_date = f"{cell_date.value.month}/{cell_date.value.day}/{cell_date.value.year}"
        else:
            fit_list_date = str(cell_date.value).strip() if cell_date.value else ""

        # 2. Convert the database dictionary keys to match
        converted_regs = {}
        for db_date, states in regs.items():
            try:
                y, m, d = db_date.split('-')
                converted_regs[f"{int(m)}/{int(d)}/{int(y)}"] = states
            except ValueError:
                converted_regs[db_date] = states

        # PRIORITY 1: Old state exists
        if old_state and old_state != "Not Found":
            cell_homestate.value = old_state

        # PRIORITY 2: Exact date match
        elif fit_list_date in converted_regs:
            cell_homestate.value = converted_regs[fit_list_date][0]
            
        # PRIORITY 3: Advisor has absolutely no registration data
        elif not regs:
            cell_homestate.value = "Not Found"
            
        # PRIORITY 4: Only one state across all registrations
        elif len(all_states) == 1: 
            print(f"{cell_fullname.value} has only one registered state: {all_states[0]}. Assigning that as home state.")
            cell_homestate.value = all_states[0]
            
        # PRIORITY 5: Multiple states, no date match -> Most Recent
        else:
            print(f"WARNING: {cell_fullname.value} has multiple states with no date match. Defaulting to most recent registration.")
            sorted_db_dates = sorted(regs.keys())
            if sorted_db_dates:
                most_recent_db_date = sorted_db_dates[-1]
                cell_homestate.value = regs[most_recent_db_date][0]

    # FIX: Safe file path splitting
    base_path, ext = os.path.splitext(fitPath)
    wb.save(get_unique_filename(f"{base_path}-TrueState{ext}"))
    
if __name__ == '__main__':
    addTrueState(r"H:\_INSTITUTIONAL DIVISION\INTERN FOLDER\Adam Neulander\IAPD_Database\7-26.xlsx", r"H:\_INSTITUTIONAL DIVISION\INTERN FOLDER\Adam Neulander\IAPD_Database\6-26-TrueState.xlsx")